#!/usr/bin/env python3
"""
import_coreomics_xlsx.py — Load Adam Olshen's coreomics submissions export
into the DE-LIMP PG database.

The coreomics system periodically exports an Excel file with two sheets:

    submissions  — one row per customer submission (61 columns)
    samples      — one row per individual sample within a submission (7 cols)

This script UPSERTs both sheets into:

    coreomics_submissions_cache
    coreomics_samples_cache

Idempotent: re-running with a fresh export merges new submissions and
updates changed ones without disturbing existing linkages in
delimp_raw_files_internal.

USAGE
-----
    # Mac (token in shared volume)
    PGPASSWORD=$(cat /Volumes/proteomics-grp/brett/.pgfarm_delimp_token) \\
    python3 scripts/import_coreomics_xlsx.py \\
        --xlsx /Users/brettphinney/Downloads/submissions_export.2026_05_19__00_29.xlsx \\
        --host pgfarm.library.ucdavis.edu \\
        --user brettsp \\
        --database uc-davis-genome-center-proteomics-core/delimp

    # Auto-link after import (one-time backfill)
    python3 scripts/import_coreomics_xlsx.py --xlsx ... --auto-link-raws

PG CONNECTION
-------------
Mirrors STAN's db_pg.py conventions:
    sslmode=require   (NOT verify-full — broken on Mac cert path)
    persistent cached connection (avoid per-row SSL handshake)
    7-day CAS bearer token from --token-file (or PGPASSWORD env)

NOTES
-----
PII filtering: submitter_email and pi_email are stored as-is (core staff
need them for follow-up). Phone numbers are dropped. This decision is
encoded in COLUMN_MAP below — adjust if the policy changes.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras


# ────────────────────────────────────────────────────────────────────────────
# Column mapping: coreomics .xlsx column → PG cache column.
# Source values are TEXT; conversions (TIMESTAMPTZ, INT) happen at insert time.
#
# Anything NOT in COLUMN_MAP stays in the raw_payload JSONB. So filtering
# out a column here doesn't lose the data — it just moves it to JSONB
# where it's queryable via `raw_payload->>'phone_number'` if ever needed.
# ────────────────────────────────────────────────────────────────────────────

COLUMN_MAP = {
    # Identity
    "ID": "submission_id",
    "Internal ID": "internal_id",
    "Type": "type",
    "Submitted": "submitted_at",       # TIMESTAMPTZ
    "Status": "status",
    "send_date": "send_date",          # DATE

    # Submitter (names + email; phone dropped per privacy policy)
    "First Name": "submitter_first_name",
    "Last Name": "submitter_last_name",
    "Submitter Email": "submitter_email",

    # PI (same policy)
    "PI First Name": "pi_first_name",
    "PI Last Name": "pi_last_name",
    "PI Email": "pi_email",

    # Institution
    "Institute": "institute",

    # Frequently-queried submission metadata
    "num_samples": "num_samples",      # INT
    "organism": "organism",
    "species": "species",
    "prot_or_pep": "prot_or_pep",
    "proteomics_type": "proteomics_type",
    "mass_spec_wanted": "mass_spec_wanted",
    "sample_prep": "sample_prep",
    "gradient_length": "gradient_length",
    "dia": "dia",
    "tmt": "tmt",
    "description": "description",
    "other_info": "other_info",

    # Compliance / billing
    "biohazard": "biohazard",
    "pathogenic": "pathogenic",
    "nih_s10_user": "nih_s10_user",
    "is_nih_major_user": "is_nih_major_user",
    "transgenic": "transgenic",
    "po_account_number": "po_account_number",
}

# Columns we explicitly DROP and don't even keep in raw_payload (true PII).
DROP_COLUMNS = {"Submitter Phone", "PI Phone"}


def md5_of_file(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def coerce(val, target_type: str):
    """Coerce an Excel cell value to the right Python type for PG."""
    if val is None or val == "":
        return None
    if target_type == "int":
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None
    if target_type == "timestamptz":
        if isinstance(val, datetime):
            return val
        try:
            return datetime.fromisoformat(str(val))
        except ValueError:
            return None
    if target_type == "date":
        if isinstance(val, datetime):
            return val.date()
        try:
            return datetime.fromisoformat(str(val)).date()
        except ValueError:
            return None
    # default: text
    return str(val).strip() if val is not None else None


COLUMN_TYPES = {
    "submitted_at": "timestamptz",
    "send_date": "date",
    "num_samples": "int",
}


def load_submissions(ws, source_export: str, source_md5: str):
    """Yield (column_dict, raw_payload_dict) per submission row."""
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    for row in rows[1:]:
        record = dict(zip(headers, row))

        # Mapped columns → PG record
        pg_record = {"source_export": source_export, "source_export_md5": source_md5}
        for src_col, pg_col in COLUMN_MAP.items():
            raw_val = record.get(src_col)
            pg_record[pg_col] = coerce(raw_val, COLUMN_TYPES.get(pg_col, "text"))

        # Everything else → JSONB payload (minus dropped PII)
        payload = {
            k: v for k, v in record.items()
            if k not in COLUMN_MAP and k not in DROP_COLUMNS and v is not None
        }
        # JSON can't serialize datetimes
        for k, v in list(payload.items()):
            if isinstance(v, datetime):
                payload[k] = v.isoformat()
        pg_record["raw_payload"] = json.dumps(payload)

        if pg_record["submission_id"]:
            yield pg_record


def load_samples(ws):
    """Yield (submission_id, unique_id, sample_name, condition_name, ...) per sample row."""
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    for row in rows[1:]:
        record = dict(zip(headers, row))
        sub_id = record.get("submission_id")
        unique_id = record.get("unique_id")
        if not (sub_id and unique_id):
            continue
        yield {
            "submission_id": str(sub_id).strip(),
            "unique_id": str(unique_id).strip(),
            "sample_name": (record.get("sample_name") or None) and str(record["sample_name"]).strip(),
            "condition_name": (record.get("condition_name") or None) and str(record["condition_name"]).strip(),
            "amt_to_inject": (record.get("amt_2_inject") or None) and str(record["amt_2_inject"]).strip(),
            "internal_id": (record.get("internal_id") or None) and str(record["internal_id"]).strip(),
            "internal_notes": (record.get("internal_notes") or None) and str(record["internal_notes"]).strip(),
        }


# ────────────────────────────────────────────────────────────────────────────
# PG side
# ────────────────────────────────────────────────────────────────────────────

UPSERT_SUBMISSION_SQL = """
INSERT INTO coreomics_submissions_cache (
    submission_id, internal_id, type, submitted_at, status, send_date,
    submitter_first_name, submitter_last_name, submitter_email,
    pi_first_name, pi_last_name, pi_email,
    institute, num_samples, organism, species, prot_or_pep, proteomics_type,
    mass_spec_wanted, sample_prep, gradient_length, dia, tmt,
    description, other_info,
    biohazard, pathogenic, nih_s10_user, is_nih_major_user, transgenic,
    po_account_number,
    raw_payload, source_export, source_export_md5
) VALUES (
    %(submission_id)s, %(internal_id)s, %(type)s, %(submitted_at)s,
    %(status)s, %(send_date)s,
    %(submitter_first_name)s, %(submitter_last_name)s, %(submitter_email)s,
    %(pi_first_name)s, %(pi_last_name)s, %(pi_email)s,
    %(institute)s, %(num_samples)s, %(organism)s, %(species)s,
    %(prot_or_pep)s, %(proteomics_type)s, %(mass_spec_wanted)s,
    %(sample_prep)s, %(gradient_length)s, %(dia)s, %(tmt)s,
    %(description)s, %(other_info)s,
    %(biohazard)s, %(pathogenic)s, %(nih_s10_user)s, %(is_nih_major_user)s,
    %(transgenic)s, %(po_account_number)s,
    %(raw_payload)s::jsonb, %(source_export)s, %(source_export_md5)s
)
ON CONFLICT (submission_id) DO UPDATE SET
    internal_id = EXCLUDED.internal_id,
    status = EXCLUDED.status,
    send_date = EXCLUDED.send_date,
    num_samples = EXCLUDED.num_samples,
    organism = EXCLUDED.organism,
    species = EXCLUDED.species,
    description = EXCLUDED.description,
    other_info = EXCLUDED.other_info,
    raw_payload = EXCLUDED.raw_payload,
    source_export = EXCLUDED.source_export,
    source_export_md5 = EXCLUDED.source_export_md5,
    imported_at = NOW();
"""

UPSERT_SAMPLE_SQL = """
INSERT INTO coreomics_samples_cache (
    submission_id, unique_id, sample_name, condition_name,
    amt_to_inject, internal_id, internal_notes
) VALUES (
    %(submission_id)s, %(unique_id)s, %(sample_name)s, %(condition_name)s,
    %(amt_to_inject)s, %(internal_id)s, %(internal_notes)s
)
ON CONFLICT (submission_id, unique_id) DO UPDATE SET
    sample_name = EXCLUDED.sample_name,
    condition_name = EXCLUDED.condition_name,
    amt_to_inject = EXCLUDED.amt_to_inject,
    internal_id = EXCLUDED.internal_id,
    internal_notes = EXCLUDED.internal_notes,
    imported_at = NOW();
"""

# Auto-link query: for any raw_file with no coreomics linkage yet, find the
# coreomics sample whose sample_name is a prefix (after normalising
# separators) of the raw basename. Picks the longest match (more specific).
AUTO_LINK_SQL = """
WITH unlinked AS (
    SELECT rf.raw_path, rf.raw_basename
    FROM raw_files rf
    LEFT JOIN delimp_raw_files_internal rfi ON rf.raw_path = rfi.raw_path
    WHERE rfi.coreomics_sample_unique_id IS NULL
       OR rfi IS NULL
),
candidates AS (
    SELECT
        u.raw_path, u.raw_basename,
        cs.submission_id, cs.unique_id, cs.sample_name, cs.condition_name,
        ROW_NUMBER() OVER (
            PARTITION BY u.raw_path
            ORDER BY LENGTH(cs.sample_name) DESC
        ) AS rank
    FROM unlinked u
    CROSS JOIN coreomics_samples_cache cs
    WHERE cs.sample_name IS NOT NULL
      AND regexp_replace(u.raw_basename, '[._-]', '', 'g')
          LIKE regexp_replace(cs.sample_name, '[._-]', '', 'g') || '%%'
)
INSERT INTO delimp_raw_files_internal (
    raw_path, coreomics_submission_id, coreomics_sample_unique_id,
    coreomics_sample_name, coreomics_condition_name, coreomics_last_synced_at
)
SELECT raw_path, submission_id, unique_id, sample_name, condition_name, NOW()
FROM candidates
WHERE rank = 1
ON CONFLICT (raw_path) DO UPDATE SET
    coreomics_submission_id = EXCLUDED.coreomics_submission_id,
    coreomics_sample_unique_id = EXCLUDED.coreomics_sample_unique_id,
    coreomics_sample_name = EXCLUDED.coreomics_sample_name,
    coreomics_condition_name = EXCLUDED.coreomics_condition_name,
    coreomics_last_synced_at = NOW()
WHERE delimp_raw_files_internal.coreomics_sample_unique_id IS NULL;
"""


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, type=Path,
                    help="Path to coreomics submissions_export_*.xlsx")
    ap.add_argument("--host", default="pgfarm.library.ucdavis.edu")
    ap.add_argument("--port", type=int, default=5432)
    ap.add_argument("--database",
                    default="uc-davis-genome-center-proteomics-core/delimp")
    ap.add_argument("--user", default=os.environ.get("USER", "brettsp"))
    ap.add_argument("--token-file",
                    help="Path to PGFarm token file (overrides PGPASSWORD env)")
    ap.add_argument("--auto-link-raws", action="store_true",
                    help="After import, run the auto-linkage pass on raw_files")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse and report counts but don't write to PG")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"FATAL: xlsx not found: {args.xlsx}")

    # ── Load .xlsx ────────────────────────────────────────────────────────
    print(f"[+] Reading {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    source_md5 = md5_of_file(args.xlsx)
    print(f"    md5: {source_md5}")
    if "submissions" not in wb.sheetnames or "samples" not in wb.sheetnames:
        sys.exit(f"FATAL: expected 'submissions' and 'samples' sheets, got {wb.sheetnames}")

    submissions = list(load_submissions(wb["submissions"], args.xlsx.name, source_md5))
    samples = list(load_samples(wb["samples"]))
    print(f"    {len(submissions)} submission rows, {len(samples)} sample rows")

    if args.dry_run:
        print("\n[dry-run] would upsert above counts. Sample submission:")
        if submissions:
            for k, v in list(submissions[0].items())[:8]:
                print(f"    {k}: {v}")
        return

    # ── PG connection ─────────────────────────────────────────────────────
    pwd = os.environ.get("PGPASSWORD")
    if args.token_file:
        pwd = Path(args.token_file).read_text().strip()
    if not pwd:
        sys.exit("FATAL: no token in PGPASSWORD env or --token-file")

    print(f"[+] Connecting to {args.host}:{args.port}/{args.database} as {args.user}")
    conn = psycopg2.connect(
        host=args.host, port=args.port,
        database=args.database, user=args.user, password=pwd,
        sslmode="require",
    )
    conn.autocommit = False

    try:
        with conn.cursor() as cur:
            # Upsert submissions (one execute_batch per chunk for speed)
            print(f"[+] Upserting {len(submissions)} submissions...")
            psycopg2.extras.execute_batch(cur, UPSERT_SUBMISSION_SQL, submissions, page_size=200)

            print(f"[+] Upserting {len(samples)} samples...")
            psycopg2.extras.execute_batch(cur, UPSERT_SAMPLE_SQL, samples, page_size=500)

            if args.auto_link_raws:
                print("[+] Running auto-link against raw_files...")
                cur.execute(AUTO_LINK_SQL)
                print(f"    linked {cur.rowcount} raw files to coreomics samples")

        conn.commit()
        print("[+] Done. Committed.")

        # Quick summary query
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM coreomics_submissions_cache;")
            n_sub = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM coreomics_samples_cache;")
            n_sam = cur.fetchone()[0]
            cur.execute(
                "SELECT COUNT(*) FROM delimp_raw_files_internal "
                "WHERE coreomics_sample_unique_id IS NOT NULL;")
            n_linked = cur.fetchone()[0]
            print(f"\nDB state after import:")
            print(f"  coreomics_submissions_cache: {n_sub:>6} rows")
            print(f"  coreomics_samples_cache:     {n_sam:>6} rows")
            print(f"  raw_files linked:            {n_linked:>6}")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
